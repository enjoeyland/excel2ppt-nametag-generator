import subprocess, os, platform

from collections import defaultdict
from openpyxl import load_workbook
from pptx.enum.dml import MSO_COLOR_TYPE, MSO_FILL
from pptx.oxml.ns import qn

class dotdict(dict):
    """dot.notation access to dictionary attributes"""
    __getattr__ = dict.get
    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__
    def copy(self):
        return dotdict(super().copy())

def chunk_list(l, chunk_size):
    return [l[i:i + chunk_size] for i in range(0, len(l), chunk_size)]

def tuples_to_dict_list(header, data):
    return [dict(zip(header, d)) for d in data]

def read_excel_data(filename):
    workbook = load_workbook(filename, data_only=True)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    assert len(data) > 0, "No data found in the excel file"
    assert len(data) > 1, "Only header found in the excel file. No data found"
    for i, row in enumerate(data):
        data[i] = tuple(str(c) if c is not None else "" for c in row)

    header = data[0]
    header = [h.strip().lower() for h in header]
        
    return header, data[1:]

def headed_data_with_sample_num(header, data):
    try:
        sample_num_idx = header.index("sample num")
    except:
        sample_num_idx = len(header)
        header += ("sample num",)
        data = [d + (0,) for d in data]
    else:
        for i, d in enumerate(data):
            d = list(d)
            if isinstance(d[sample_num_idx], str):
                if d[sample_num_idx].strip() == "":
                    continue
                try:
                    d[sample_num_idx] = int(d[sample_num_idx].strip())
                except:
                    raise ValueError(f"Sample number '{d[sample_num_idx]}' is not an integer in row {i+2}")
            data[i] = tuple(d)
    data = tuples_to_dict_list(header, data)
    return data

def group_by_sample(data):
    data_by_sample = defaultdict(list)
    for d in data:
        data_by_sample[d["sample num"]].append(d)
    return data_by_sample

def get_data_by_sample(filename):
    header, data = read_excel_data(filename)
    haeded_data = headed_data_with_sample_num(header, data)
    return group_by_sample(haeded_data)

def open_file_with_default_program(filename):
    if os.path.isfile(filename):
        if platform.system() == 'Windows':
            os.startfile(filename)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(('open', filename))
        else:  # linux variants
            subprocess.run(('xdg-open', filename))
    else:
        raise FileNotFoundError(f"No file found at {filename}")
    
def set_fill(source_shape, target_shape):
    try:
        if source_shape.fill.type is None:
            pass
        elif source_shape.fill.type == MSO_FILL.BACKGROUND:
            target_shape.fill.background()
        elif source_shape.fill.type == MSO_FILL.SOLID:
            if source_shape.fill.fore_color.type == MSO_COLOR_TYPE.SCHEME:
                color = source_shape.fill.fore_color.theme_color
                target_shape.fill.solid()
                target_shape.fill.fore_color.theme_color = color
                target_shape.fill.fore_color.brightness = source_shape.fill.fore_color.brightness
            elif source_shape.fill.fore_color.type == MSO_COLOR_TYPE.RGB:
                color = source_shape.fill.fore_color.rgb
                target_shape.fill.solid()
                target_shape.fill.fore_color.rgb = color
            elif source_shape.fill.fore_color.type is None:
                target_shape.fill.background()
            else:
                print(f"Unsupported color type: {source_shape.fill.fore_color.type}")
        elif source_shape.fill.type == MSO_FILL.PICTURE:
            import xml.etree.ElementTree as ET
            from pptx.oxml import parse_xml
            from io import BytesIO
            
            source_blipFill = source_shape.fill._xPr.find(qn('a:blipFill'))
            blip_elem = source_blipFill.find(qn('a:blip'))
            old_rId = blip_elem.get(qn('r:embed'))

            # 원본 이미지 가져오기
            source_image = source_shape.part.get_image(old_rId)
            image_blob = BytesIO(source_image.blob)
            
            # target slide에 이미지 추가
            target_slide_shapes = target_shape.part.slide.shapes
            image_part, new_rId = target_slide_shapes.part.get_or_add_image_part(image_blob)
            
            # blipFill XML 복사
            blipFill_xml = ET.tostring(source_blipFill, encoding='unicode')
            blipFill_copy = parse_xml(blipFill_xml)
            
            # rId 업데이트
            blip_copy_elem = blipFill_copy.find(qn('a:blip'))
            if blip_copy_elem is not None:
                blip_copy_elem.set(qn('r:embed'), new_rId)

            # 새 blipFill 추가
            target_shape.fill._xPr.append(blipFill_copy)
        else:
            print(f"Unsupported fill type: {source_shape.fill.type}")
    except TypeError:
        pass

def set_line(source_line, target_line): # _BasePicture, Shape, Connector
    target_line.dash_style = source_line.dash_style
    target_line.width = source_line.width
    set_fill(source_line, target_line)

def set_shadow(source_shadow, target_shadow): # BaseShape
    if not source_shadow.inherit:
        ...
        # target_shadow.blur_radius = source_shadow.blur_radius
        # target_shadow.distance = source_shadow.distance
        # target_shadow.direction = source_shadow.direction
        # target_shadow.rotation = source_shadow.rotation
        # set_color(source_shadow.color, target_shadow.color)
        # target_shadow.visible = source_shadow.visible

def set_base_shape(source_shape, target_shape): # BaseShape
    target_shape.rotation = source_shape.rotation
    set_shadow(source_shape.shadow, target_shape.shadow)


def qn_xpath(xpath: str) -> str:
    """
    XPath 문자열에서 네임스페이스 접두어(prefix)를 자동 변환하여 Clark-notation 형식으로 변환하는 함수.
    
    :param xpath: 네임스페이스 접두어(prefix)가 포함된 XPath 문자열 (예: ".//p:guideLst")
    :return: 변환된 XPath 문자열 (예: ".//{http://schemas...}guideLst")
    """
    parts = xpath.split('/')  # XPath를 '/' 기준으로 나누기
    transformed_parts = [qn(tag) if ':' in tag else tag for tag in parts]  # 네임스페이스가 있는 경우 변환
    return '/'.join(transformed_parts)  # 변환된 태그들을 다시 합쳐서 반환

def pretty_print_xml(element):
    """XML을 보기 좋게 출력하는 함수"""
    import xml.etree.ElementTree as ET
    import xml.dom.minidom
    rough_string = ET.tostring(element, encoding="unicode")
    reparsed = xml.dom.minidom.parseString(rough_string)
    print(reparsed.toprettyxml(indent="  "))